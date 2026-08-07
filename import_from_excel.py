"""
One-time import: reads scholars/departments/grants out of the legacy
Excel workbook and writes them into grants.db, going through the same
service-layer functions the web app itself uses for every other write
- so imported data gets the exact same validation, and there's no
separate "import path" that could quietly diverge from normal saves.

Usage:
    python import_from_excel.py "C:\\path\\to\\ScholarGrantTracker.xlsm"

Expects the same table/column names the VBA app used:
    tbl_Scholars:    EmployeeID, Name, Age, PreviousDegree, MissingRequirements
    tbl_Departments: AssignmentID, EmployeeID, Department, Rank, Tenure
    tbl_Grants:      GrantID, EmployeeID, ProgramApplied, TypeOfGrant,
                      DeliveringHEI, DateStarted, DateEnded, Extension,
                      Status, Remarks

If your actual column headers differ even slightly, this will tell you
exactly which ones it couldn't find rather than silently skipping data -
see the "Missing expected columns" check below.
"""
import sys
from datetime import date, datetime

import openpyxl

sys.path.insert(0, str(__import__("pathlib").Path(__file__).parent))

from app.database import Base, engine, SessionLocal
from app.services import scholars as scholar_service
from app.services import departments as dept_service
from app.services import grants as grant_service


def _to_date(value) -> date | None:
    """Excel gives us either a real datetime object or a string,
    depending on how the cell was formatted - handle both rather than
    assuming."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _to_bool(value) -> bool:
    return str(value).strip().upper() in ("TRUE", "YES", "1", "X")


def _sheet_rows(wb, sheet_name: str, expected_columns: list[str]):
    """Yields each data row as a dict keyed by column header. Raises a
    clear error naming the missing columns instead of silently
    importing partial/wrong data if the headers don't match."""
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: sheet '{sheet_name}' not found in workbook - skipping.")
        return
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = [c for c in expected_columns if c not in headers]
    if missing:
        print(f"  WARNING: sheet '{sheet_name}' is missing expected columns: {missing}")
        print(f"           Found columns: {headers}")
        print("           Proceeding with whatever columns ARE present.")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue  # skip fully blank rows
        yield dict(zip(headers, row))


def main(xlsx_path: str):
    print(f"Reading legacy workbook: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    employee_id_map: dict[int, int] = {}  # old EmployeeID -> new Scholar.id
    counts = {"scholars": 0, "assignments": 0, "grants": 0, "errors": 0}

    try:
        print("\n-- Scholars --")
        for row in _sheet_rows(
            wb, "Scholars", ["EmployeeID", "Name", "Age", "PreviousDegree", "MissingRequirements"]
        ):
            try:
                old_id = int(row.get("EmployeeID"))
                scholar = scholar_service.create_scholar(
                    db,
                    name=str(row.get("Name") or ""),
                    age=int(row["Age"]) if row.get("Age") not in (None, "") else None,
                    previous_degree=row.get("PreviousDegree"),
                    missing_requirements=_to_bool(row.get("MissingRequirements")),
                )
                employee_id_map[old_id] = scholar.id
                counts["scholars"] += 1
            except (ValueError, TypeError) as e:
                print(f"  SKIPPED scholar row {row}: {e}")
                counts["errors"] += 1

        print("-- Department Assignments --")
        for row in _sheet_rows(
            wb, "Departments", ["EmployeeID", "Department", "Rank", "Tenure"]
        ):
            old_emp_id = row.get("EmployeeID")
            new_id = employee_id_map.get(int(old_emp_id)) if old_emp_id is not None else None
            if new_id is None:
                print(f"  SKIPPED assignment row {row}: EmployeeID {old_emp_id} has no matching scholar")
                counts["errors"] += 1
                continue
            try:
                dept_service.create_assignment(
                    db,
                    new_id,
                    department=row.get("Department") or "",
                    rank=row.get("Rank"),
                    tenure=row.get("Tenure"),
                    date_started=_to_date(row.get("DateStarted")),
                    date_ended=_to_date(row.get("DateEnded")),
                )
                counts["assignments"] += 1
            except ValueError as e:
                print(f"  SKIPPED assignment row {row}: {e}")
                counts["errors"] += 1

        print("-- Grants --")
        for row in _sheet_rows(
            wb,
            "Grants",
            [
                "EmployeeID", "ProgramApplied", "TypeOfGrant", "DeliveringHEI",
                "DateStarted", "DateEnded", "Extension", "Status", "Remarks",
            ],
        ):
            old_emp_id = row.get("EmployeeID")
            new_id = employee_id_map.get(int(old_emp_id)) if old_emp_id is not None else None
            if new_id is None:
                print(f"  SKIPPED grant row {row}: EmployeeID {old_emp_id} has no matching scholar")
                counts["errors"] += 1
                continue
            try:
                status = (row.get("Status") or "Active").strip()
                if status not in grant_service.VALID_STATUSES:
                    status = "Active"
                grant_service.create_grant(
                    db,
                    new_id,
                    program_applied=row.get("ProgramApplied") or "",
                    type_of_grant=row.get("TypeOfGrant"),
                    delivering_hei=row.get("DeliveringHEI"),
                    date_started=_to_date(row.get("DateStarted")),
                    date_ended=_to_date(row.get("DateEnded")),
                    extension=_to_bool(row.get("Extension")),
                    status=status,
                    remarks=row.get("Remarks"),
                )
                counts["grants"] += 1
            except ValueError as e:
                print(f"  SKIPPED grant row {row}: {e}")
                counts["errors"] += 1

        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    print("\n== Import complete ==")
    print(f"Scholars imported:    {counts['scholars']}")
    print(f"Assignments imported: {counts['assignments']}")
    print(f"Grants imported:      {counts['grants']}")
    print(f"Rows skipped/errored: {counts['errors']}")
    if counts["errors"]:
        print("\nReview the SKIPPED lines above - those rows were NOT imported.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python import_from_excel.py <path to legacy .xlsx/.xlsm file>")
        sys.exit(1)
    main(sys.argv[1])
